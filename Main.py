import openpyxl as op
import os
from Automator import Automator

#Below are Notes for the openpyxl module -- may be able ot use for allocations prjt

'''
List of useful commands:

op.__version__

type(workbook)

 print(wb.sheetnames)

    sheet = wb['Notes1']

    print(type(sheet))

    print(sheet['A1'].value)

    print(sheet['B1'].value)

    print(sheet['C1'].value)

    #sheet['D1'] = "Hi my name is Sameet Hegde!"
   # wb.save('Practice_openpyxl.xlsx') #can also be used to save in different files

    print(sheet.cell(row=1, column=2).value) #another way ot do this

    print(sheet.max_row)
    print(sheet.max_column)

    print(wb.sheetnames[1])
    print(type(wb.sheetnames[1]))


    if(all(sheetname != "createSheet2" for sheetname in wb.sheetnames)):
        wb.create_sheet(title="createSheet2")
        wb.save("Practice_openpyxl.xlsx")
        
    
    folder = os.path.expanduser("~/Desktop/Allocations_Submissions") #os.path.expanduser allows us to use shell path
    os.chdir(folder)
    wb = op.load_workbook('submissions.xlsx')

    print(wb.sheetnames)
    sheet = wb['submissions']

    print(sheet['A1'].value)

    rowlist = list(sheet.rows)


    newWb = op.Workbook()

    ws1 = newWb.create_sheet("Test")

    for cell in rowlist[2]:
        ws1.append()



    newWb.save("Test.xlsx")


'''

if __name__ == "__main__":

    for i in range(3,15,2):
        automateTest = Automator(i)
        automateTest.create_Page()
        automateTest.populate_page()









