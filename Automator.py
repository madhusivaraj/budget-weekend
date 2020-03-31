import openpyxl as op
import os
import datetime

class Automator:

    eventCount = 0
    eventList = {}
    moveCells_OrgMaintenance = {'B23': 'X', 'B24': 'AC3', 'B26': 'AA', 'B27': 'Y', 'B28': 'AB', 'B30': 'W',
                                'B33': 'Z', 'B34': 'AE', 'B36': 'AD', 'B37': 'AF', 'B38': 'AG'}

    #Programming columns
    #Program Name, Event Description, location, Attendance, Date, Room Rent and Equip., Advertising, Beverage, Supplies/Decorations, Duplications, Contracts(not included yet), Other,
    moveCells_ProgrammingBool = False
    moveCells_Programming = {'A2': ['BR', 'FU', 'LC'], 'A3': ['BS', 'FV', 'LD'], 'B3': ['BV', 'FY', 'LG'], 'C3': ['BU', 'FX', 'LF'],
                 'D3': ['BT', 'FW', 'LE'], 'B5': ['BY', 'GB', 'LJ'], 'B6': ['BZ', 'GC', 'LK'], 'B7': ['CA', 'GD', 'LL'],
                 'B8': ['CB', 'GE', 'LM'], 'B9': ['CC', 'GF', 'LN'], 'B12':['CK','GN','LV'],  'B16': ['CL', 'GO', 'LW'], 'D5': ['CO', 'GQ', 'LZ'],
                 'D6': ['CP', 'GR', 'MA'], 'D7': ['CQ', 'GS', 'MB'], 'D8': ['CR', 'GT', 'MC'],'D9': ['CS', 'GU', 'MD'], 'D12':['CT','GV','ME'],
                 'D16': ['CK', 'GW', 'MF']
                 }

    moveCells_ProgrammingBool2 = False
    moveCells_Programming2 = {'G2': ['BR', 'FU', 'LC'], 'G3': ['BS', 'FV', 'LD'], 'H3': ['BV', 'FY', 'LG'], 'I3': ['BU', 'FX', 'LF'],
                  'J3': ['BT', 'FW', 'LE'], 'H5': ['BY', 'GB', 'LJ'], 'H6': ['BZ', 'GC', 'LK'], 'H7': ['CA', 'GD', 'LL'],
                  'H8': ['CB', 'GE', 'LM'], 'H9': ['CC', 'GF', 'LN'], 'H12':['CK','GN','LV'], 'H16': ['CL', 'GO', 'LW'], 'J5': ['CO', 'GQ', 'LZ'],
                  'J6': ['CP', 'GR', 'MA'], 'J7': ['CQ', 'GS', 'MB'], 'J8': ['CR', 'GT', 'MC'], 'J9': ['CS', 'GU', 'MD'], 'J12':['CT','GV','ME'],
                  'J16': ['CK', 'GW', 'MF'],
                  }

    #Series Programming Column
    moveCells_seriesProgrammingBool = False
    moveCells_seriesProgramming = {'H20':['CW','HA','LC'], 'H21':['CX', 'HB', ''], 'L20':['CY', 'GZ', 'LB'], 'L21':['DA','HD','LF'],
                                   'H23':['DE','HH','LJ'], 'H26':['DF', 'HI','LK'], 'H27':['DG', 'HJ', 'LL'], 'H30':['DH','HK','LM'],
                                   'H31':['DI','HL','LN'], 'H34':['DQ','HT','LV'], 'H37':['DR','HU','LW'], 'J23':['DU', 'HX', 'LZ'],
                                   'J26':['DV','HY','MA'], 'J27':['DW','HZ','MB'], 'J30':['DX','IA','MC'], 'J31':['DY','IB','MD'],
                                   'J34':['DC','IC','ME'],'J37':['EA','ID','MF']
                                   }

    #Trip Competition/Conference
    moveCells_TripsCC1Bool = False
    moveCells_tripsCC1= {'A52': ['EC', 'IF', 'MH'],  'A54': ['EF', 'II', 'MK'], 'C54': ['ED', 'IG', 'MI'], 'D54': ['EE', 'IH', 'MJ'],
                         'E54': ['EG', 'IJ', 'ML'], 'F54': ['EH', 'IK', 'MM'], 'B56': ['EJ', 'IM', 'MO'], 'B57': ['EK', 'IN', 'MP'], 'B58': ['EL', 'IO', 'MQ'],
                        'B59': ['EM', 'IP', 'MR'], 'B60': ['EN', 'IQ', 'MS'], 'B61': ['EO', 'IR', 'MT'], 'D56': ['ER', 'IU', 'MW'],
                        'D57': ['ES', 'IV', 'MX'], 'D58': ['ET', 'IW', 'MY'], 'D59': ['EU', 'IX', 'MZ'], 'D60': ['EV', 'IY', 'NA'],
                        'D61': ['EW', 'IZ', 'NB'],
                        }



    moveCells_TripsCC2Bool = False
    moveCells_TripsCC2 = {'G52': ['EC', 'IF', 'MH'],  'G54': ['EF', 'II', 'MK'], 'I54': ['ED', 'IG', 'MI'], 'J54': ['EE', 'IH', 'MJ'],
                         'K54': ['EG', 'IJ', 'ML'], 'L54': ['EH', 'IK', 'MM'], 'H56': ['EJ', 'IM', 'MO'], 'H57': ['EK', 'IN', 'MP'], 'H58': ['EL', 'IO', 'MQ'],
                        'H59': ['EM', 'IP', 'MR'], 'H60': ['EN', 'IQ', 'MS'], 'H61': ['EO', 'IR', 'MT'], 'J56': ['ER', 'IU', 'MW'],
                        'J57': ['ES', 'IV', 'MX'], 'J58': ['ET', 'IW', 'MY'], 'J59': ['EU', 'IX', 'MZ'], 'J60': ['EV', 'IY', 'NA'],
                        'J61': ['EW', 'IZ', 'NB'],
                        }


    #Other Trip
    moveCells_tripsOtherBool = False
    moveCells_otherTrip = {'I40': ['EY', 'JA', 'ND'], 'H41': ['EY', 'JA', 'ND'], 'H42': ['FB', 'JD', 'NG'],
                           'H44': ['FG', 'JI', 'NL'],
                           'H45': ['FF', 'JH', 'NK'], 'H46': ['FH', 'JJ', 'NM'], 'H47': ['FI', 'JK', 'NN'],
                           'H48': ['FJ', 'JL', 'NO'],
                           'H49': ['FK', 'JM', 'NP'], 'J44': ['FO', 'JQ', 'NT'], 'J45': ['FN', 'JP', 'NS'],
                           'J46': ['FP', 'JR', 'NU'],
                           'J47': ['FQ', 'JS', 'NV'], 'J48': ['FR', 'JT', 'NW'], 'J49': ['FS', 'JU', 'NX'],
                           }

    def __init__(self, row, folder=os.path.expanduser("~/Desktop/Allocations_Submissions")):
        self.folder = folder
        os.chdir(folder)
        self.capSheet = op.load_workbook("capSheet.xlsx")
        self.submissions = op.load_workbook("submissions.xlsx")
        self.submissionSheet = self.submissions.active
        self.target = None
        self.row = row


    def create_Page(self):

        print("in create_Page")
        #newWb = op.Workbook()

        cap = self.capSheet.active

        self.target = self.capSheet.copy_worksheet(cap)

        self.target.title = self.submissionSheet['G' + str(self.row)].value
        self.target['A1'] = self.target['A1'].value + self.submissionSheet['G' + str(self.row)].value
        self.target['K1'] = self.target['K1'].value + ' ' + str(self.submissionSheet['H' + str(self.row)].value)
        print(self.target['K1'].value)



    def populate_page(self):
        print("in populate page")
        print(self.capSheet.sheetnames)

        sheet = self.target

        #print(sheet['A1'].value)

        #Programs: BR, CW, EC, EY, FU, HA, IF, JA, JW, LC, MH, ND
        eventListColumns = {'BR' : 'program', 'CW':'programSeries', 'EC': 'tripCC', 'EY': 'tripOther',
                            'FU': 'program', 'HA': 'programSeries', 'IF': 'tripCC', 'JA': 'tripOther', 'JW': 'program',
                            'LC': 'programSeries', 'MH':'tripCC', 'ND':'tripOther'}

        eventListRow = {key + str(self.row): value for key, value in eventListColumns.items()}
        print(eventListRow)


        for eventCell, type in eventListRow.items():
            if self.submissionSheet[eventCell].value is not None:
                self.eventList.update({eventCell:type})
                self.eventCount+=1


        print(self.eventList)
        self.populate_orgMaintenance()


        #Still need to account for >2 events
        for eventCell, type in self.eventList.copy().items():
            if type == 'program':
                self.populate_programStandAlone(eventCell)
            elif type == 'programSeries':
                self.populate_programSeries(eventCell)
            elif type == 'tripCC':
                self.populate_tripsCC(eventCell)
            elif type == 'tripOther':
                self.populate_tripsOther(eventCell)

            self.eventCount-=1
            del self.eventList[eventCell]


        self.capSheet.save("capSheet.xlsx")


    def populate_orgMaintenance(self):

        print("in populate_orgMaintenance")
        sheet = self.target


        for capColumn,submissionCell in self.moveCells_OrgMaintenance.items():
            if self.submissionSheet[submissionCell] is not None:
                row = str(self.row)
                sheet[capColumn] = self.submissionSheet[submissionCell + row].value
            else:
                continue

        print("finished orgMaintenance")

    #TODO: Still need to do Contracts and Rights
    def populate_programStandAlone(self,eventCell):

        print("in populate_ProgrammingStandAlone")

        sheet = self.target

        row = str(self.row)

        if self.moveCells_ProgrammingBool is False:

            firstKey = next(iter(self.moveCells_Programming))
            li = self.moveCells_Programming[firstKey]
            liRow = [item + str(self.row) for item in li]
            print(liRow)
            index = liRow.index(eventCell)

            for capColumn, submissionCell in self.moveCells_Programming.items():
                submissionCell = submissionCell[index]
                if self.submissionSheet[submissionCell + row] is not None:
                    sheet[capColumn] = self.submissionSheet[submissionCell + row].value

                else:
                    continue
            self.moveCells_ProgrammingBool= True

        elif self.moveCells_ProgrammingBool2 is False:

            firstKey = next(iter(self.moveCells_Programming))
            li = self.moveCells_Programming[firstKey]
            liRow = [item + str(self.row) for item in li]
            print(liRow)
            index = liRow.index(eventCell)

            for capColumn, submissionCell in self.moveCells_Programming2.items():
                submissionCell = submissionCell[index]
                if self.submissionSheet[submissionCell + row] is not None:
                   # print(self.submissionSheet[submissionCell + row].value)
                    sheet[capColumn] = self.submissionSheet[submissionCell + row].value

                else:
                    continue

            self.moveCells_ProgrammingBool2 = True



    def populate_programSeries(self, eventCell):
        print("in populateProgrammingSeries")

        sheet = self.target
        row = str(self.row)

        if self.moveCells_seriesProgrammingBool is False:

            firstKey = next(iter(self.moveCells_seriesProgramming))
            li = self.moveCells_seriesProgramming[firstKey]
            liRow = [item + str(self.row) for item in li]
            index = liRow.index(eventCell)

            for capColumn, submissionCell in self.moveCells_seriesProgramming.items():
                submissionCell = submissionCell[index]
                if self.submissionSheet[submissionCell + row] is not None:
                    cell = submissionCell + row
                    sheet[capColumn] = self.submissionSheet[cell].value

                else:
                    continue

            self.moveCells_seriesProgrammingBool = True

    def populate_tripsCC(self, eventCell):
        print("in populate_tripsCC")

        sheet = self.target

        row = str(self.row)

        if self.moveCells_TripsCC1Bool is False:

            firstKey = next(iter(self.moveCells_tripsCC1))
            li = self.moveCells_tripsCC1[firstKey]
            liRow = [item + str(self.row) for item in li]
            print(liRow)
            index = liRow.index(eventCell)

            for capColumn, submissionCell in self.moveCells_tripsCC1.items():
                submissionCell = submissionCell[index]
                if self.submissionSheet[submissionCell + row] is not None:
                    sheet[capColumn] = self.submissionSheet[submissionCell + row].value

                else:
                    continue
            self.moveCells_TripsCC1Bool = True

        elif self.moveCells_TripsCC2Bool is False:

            firstKey = next(iter(self.moveCells_TripsCC2))
            li = self.moveCells_tripsCC2[firstKey]
            liRow = [item + str(self.row) for item in li]
            print(liRow)
            index = liRow.index(eventCell)

            for capColumn, submissionCell in self.moveCells_TripsCC2.items():
                submissionCell = submissionCell[index]
                if self.submissionSheet[submissionCell + row] is not None:
                    # print(self.submissionSheet[submissionCell + row].value)
                    sheet[capColumn] = self.submissionSheet[submissionCell + row].value

                else:
                    continue

            self.moveCells_TripsCC2Bool = True

    def populate_tripsOther(self, eventCell):
        print("in populate_tripsOther")

        sheet = self.target
        row = str(self.row)

        if self.moveCells_tripsOtherBool is False:

            firstKey = next(iter(self.moveCells_otherTrip))
            li = self.moveCells_otherTrip[firstKey]
            liRow = [item + str(self.row) for item in li]
            index = liRow.index(eventCell)

            for capColumn, submissionCell in self.moveCells_otherTrip.items():
                submissionCell = submissionCell[index]
                if self.submissionSheet[submissionCell + row] is not None:
                    cell = submissionCell + row
                    sheet[capColumn] = self.submissionSheet[cell].value

                else:
                    continue

            self.moveCells_tripsOtherBool = True




'''
      if isinstance(self.submissionSheet[submissionCell + row].value, datetime.date):
                            item = datetime.datetime.strftime(self.submissionSheet[submissionCell + row].value, datetime.date)
                        sheet[capColumn] = sheet[capColumn].value + str(item)
                        
                        if sheet[capColumn].value is not None:
                        item = self.submissionSheet[submissionCell + row].value
                  

                    else: sheet[capColumn] + self.submissionSheet[submissionCell + row].value


'''